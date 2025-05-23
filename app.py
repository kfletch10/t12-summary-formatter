# T12 Formatter Streamlit App
# Author: OpenAI ChatGPT
# Description: Web-based T12 Excel formatter for multifamily real estate reports

import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile

def format_t12(file_path):
    # Load original workbook for reference values
    original_wb = load_workbook(file_path, data_only=True)
    original_ws = original_wb.active

    # Extract C9:N9 and B11:N11 values and formats
    row9_values = [original_ws.cell(row=9, column=col).value for col in range(3, 15)]
    row9_format = [original_ws.cell(row=9, column=col).number_format for col in range(3, 15)]
    row11_values = [original_ws.cell(row=11, column=col).value for col in range(2, 15)]
    row11_format = [original_ws.cell(row=11, column=col).number_format for col in range(2, 15)]

    # Reload workbook for formatting
    wb = load_workbook(file_path)
    ws = wb.active

    # Step 1: Remove unnecessary rows
    rows_to_delete = [1, 2, 3, 4, 5, 9, 10, 11, 60]
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)

    # Step 2: Unmerge A1:N49
    for merged_range in list(ws.merged_cells.ranges):
        min_row, max_row = merged_range.min_row, merged_range.max_row
        min_col, max_col = merged_range.min_col, merged_range.max_col
        if min_row <= 49 and min_col <= 14:
            ws.unmerge_cells(str(merged_range))

    # Step 3: Align left A1:A3
    for row in range(1, 4):
        ws[f"A{row}"].alignment = Alignment(horizontal='left')

    # Step 4: Set column widths B–N
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Step 5: Freeze Pane at B6
    ws.freeze_panes = "B6"

    # Step 6: Shade row 37
    fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col in range(1, 15):
        ws.cell(row=37, column=col).fill = fill

    # Step 7: Bold text in key rows
    bold_font = Font(bold=True)
    for row in [13, 17, 35, 37, 47]:
        for col in range(1, 15):
            ws.cell(row=row, column=col).font = bold_font

    # Step 8: Row height adjustments
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    # Step 9: Add and align 'Total' in N5
    ws["N5"].value = "Total"
    ws["N5"].alignment = Alignment(horizontal='right')

    # Step 10: Hide gridlines
    ws.sheet_view.showGridLines = False

    # Step 11: Restore extracted values to rows 9 and 11 with format
    for i, col in enumerate(range(3, 15)):
        ws.cell(row=9, column=col).value = row9_values[i]
        ws.cell(row=9, column=col).number_format = row9_format[i]

    for i, col in enumerate(range(2, 15)):
        ws.cell(row=11, column=col).value = row11_values[i]
        ws.cell(row=11, column=col).number_format = row11_format[i]
        ws.cell(row=11, column=col).font = Font(color="000000")

    # Step 12: Build output filename
    property_name = ws["A1"].value or "Property"
    date_value = ws["A3"].value or "NoDate"
    try:
        parsed_date = datetime.strptime(str(date_value), "%B %d, %Y")
        formatted_date = parsed_date.strftime("%Y-%m")
    except ValueError:
        formatted_date = "Unknown_Date"

    safe_property = str(property_name).replace(" ", "_").replace("/", "-").strip()
    filename = f"{safe_property}_T12_{formatted_date}.xlsx"
    output_path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(output_path)

    return output_path

# Streamlit App UI
st.title("📊 T12 Formatter for Multifamily Reports")
st.write("Upload a raw T12 Excel file, and download the investor-ready version in seconds.")

uploaded_file = st.file_uploader("Upload T12 Excel File", type=["xlsx"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        formatted_path = format_t12(tmp.name)

    with open(formatted_path, "rb") as f:
        st.download_button(
            label="📥 Download Formatted T12",
            data=f,
            file_name=os.path.basename(formatted_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
